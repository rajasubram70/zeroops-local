"""
ZeroOps Customer JSON Generator
================================
Fill in the ZeroOps_Contextualisation_Template.xlsx workbook,
then run this script to generate all 13 JSON files for the new customer.

Usage:
    python zeroops_generator.py --workbook Twinings_ZeroOps.xlsx

Output:
    src/data/customer/<customer_name>/  (13 JSON files)

Then in loader.js set:
    const ACTIVE_CUSTOMER = "<customer_name>"
"""

import json, re, sys, os, argparse, random
from datetime import datetime, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Run: pip install openpyxl --break-system-packages")
    sys.exit(1)

# ── Colours ───────────────────────────────────────────────────
NAVY   = "0B1629"
TEAL   = "0E7C7B"
BLUE   = "1D4ED8"
GREEN  = "16A34A"
AMBER  = "D97706"
RED    = "DC2626"
PURP   = "7C3AED"
WHITE  = "FFFFFF"
LGREY  = "F1F5F9"
MGREY  = "CBD5E1"

def hdr(wb_color, font_color=WHITE, bold=True, size=10):
    return {
        "fill": PatternFill("solid", fgColor=wb_color),
        "font": Font(bold=bold, color=font_color, size=size, name="Calibri"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }

def cell_style(fill=None, bold=False, color="1E293B", size=9, wrap=True, align="left"):
    s = {
        "font": Font(bold=bold, color=color, size=size, name="Calibri"),
        "alignment": Alignment(horizontal=align, vertical="center", wrap_text=wrap),
    }
    if fill:
        s["fill"] = PatternFill("solid", fgColor=fill)
    return s

def apply(cell, style):
    for k, v in style.items():
        setattr(cell, k, v)

def border_all(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color="CBD5E1")
    b = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                             min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = b

def add_sheet_header(ws, title, subtitle, color=NAVY):
    ws.merge_cells("A1:Z1")
    ws["A1"] = title
    apply(ws["A1"], hdr(color, WHITE, True, 13))
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:Z2")
    ws["A2"] = subtitle
    apply(ws["A2"], {"font": Font(italic=True, color="64748B", size=9, name="Calibri"),
                      "alignment": Alignment(horizontal="left", vertical="center")})
    ws.row_dimensions[2].height = 18

def add_col_headers(ws, row, headers, colors):
    for j, (h, c) in enumerate(zip(headers, colors), 1):
        cell = ws.cell(row=row, column=j, value=h)
        apply(cell, hdr(c, WHITE, True, 9))
    ws.row_dimensions[row].height = 22

def add_example_row(ws, row, values, shade=False):
    bg = LGREY if shade else WHITE
    for j, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=j, value=v)
        apply(cell, cell_style(fill=bg, size=9))
    ws.row_dimensions[row].height = 16

# ══════════════════════════════════════════════════════════════
# TEMPLATE BUILDER
# ══════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════
# JSON GENERATOR
# ══════════════════════════════════════════════════════════════

def read_sheet_as_list(ws, header_row=3):
    """Read a sheet into list of dicts using row header_row as keys."""
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column+1)]
    headers = [str(h).strip() if h else f"col{c}" for c, h in enumerate(headers, 1)]
    rows = []
    for r in range(header_row+1, ws.max_row+1):
        row = {}
        for c, h in enumerate(headers, 1):
            v = ws.cell(r, c).value
            row[h] = str(v).strip() if v is not None else ""
        if any(row.values()):
            rows.append(row)
    return rows

def read_profile(ws):
    """Read key-value profile sheet."""
    d = {}
    for r in range(4, ws.max_row+1):
        k = ws.cell(r, 1).value
        v = ws.cell(r, 2).value
        if k:
            d[str(k).strip()] = str(v).strip() if v is not None else ""
    return d

def safe_float(v, default=0):
    try: return float(str(v).replace("%","").replace("$","").replace(",",""))
    except: return default

def safe_int(v, default=0):
    try: return int(safe_float(v))
    except: return default

def generate_jsons(workbook_path):
    wb = openpyxl.load_workbook(workbook_path)

    def get_sheet(name_fragment):
        for name in wb.sheetnames:
            if name_fragment.lower() in name.lower():
                return wb[name]
        return None

    # Read all sheets
    profile_ws  = get_sheet("profile")
    metrics_ws  = get_sheet("kpi")
    apps_ws     = get_sheet("application")
    inc_ws      = get_sheet("incident")
    agents_ws   = get_sheet("agent")
    scen_ws     = get_sheet("scenario")
    req_ws      = get_sheet("request")
    mim_ws      = get_sheet("mim")

    if not profile_ws:
        print("❌ Could not find 'Customer Profile' sheet"); return

    profile  = read_profile(profile_ws)
    metrics  = read_sheet_as_list(metrics_ws) if metrics_ws else []
    apps     = read_sheet_as_list(apps_ws)    if apps_ws    else []
    incs     = read_sheet_as_list(inc_ws)     if inc_ws     else []
    agents   = read_sheet_as_list(agents_ws)  if agents_ws  else []
    scens    = read_sheet_as_list(scen_ws)    if scen_ws    else []
    reqs     = read_sheet_as_list(req_ws)     if req_ws     else []
    mim      = read_sheet_as_list(mim_ws)     if mim_ws     else []

    # ── KPI dict ──────────────────────────────────────────────
    kpi = {row["METRIC KEY"]: row["BASELINE VALUE"]
           for row in metrics if row.get("METRIC KEY")}
    for row in metrics:
        if row.get("CURRENT VALUE"):
            kpi[row["METRIC KEY"]] = row["CURRENT VALUE"]

    cname = profile.get("customer_name","Customer")
    slug  = re.sub(r"[^a-z0-9]","_", cname.lower()).strip("_")
    out   = Path(f"src/data/customer/{slug}")
    out.mkdir(parents=True, exist_ok=True)
    print(f"\n📁 Output directory: {out}")

    # ── 1. customer.json ──────────────────────────────────────
    customer_json = {
        "name":         cname,
        "shortName":    profile.get("short_name", cname),
        "logo":         profile.get("logo","🏢"),
        "industry":     profile.get("industry",""),
        "region":       profile.get("region",""),
        "primaryCloud": profile.get("primary_cloud","Cloud"),
        "itsm":         profile.get("itsm","ITSM"),
        "monitoring":   profile.get("monitoring","ELK Stack"),
        "automation":   profile.get("automation","Custom Scripts"),
        "incumbent":    profile.get("incumbent",""),
        "users":        profile.get("users",""),
        "accentColor":  profile.get("accent_color","#1D4ED8"),
        "transferConfig": {
            "currentMonth": safe_int(profile.get("fast_forward_months","6")),
            "overallReadiness": 64,
            "tcsShare": safe_int(profile.get("gcc_split_tcs","70")),
            "gccShare": safe_int(profile.get("gcc_split_gcc","30")),
            "handoverDate": profile.get("handover_date","Jun 2027"),
        }
    }
    (out/"customer.json").write_text(json.dumps(customer_json, indent=2))
    print("  ✅ customer.json")

    # ── 2. credentials.json ───────────────────────────────────
    creds = {"credentials":[{
        "username": profile.get("demo_login_user", f"aiops.engineer@{slug}.com"),
        "password": profile.get("demo_login_pass","ZeroOps2026"),
        "role": "aiops_engineer", "name":"AIOps Engineer"
    }]}
    (out/"credentials.json").write_text(json.dumps(creds, indent=2))
    print("  ✅ credentials.json")

    # ── 3. roles.json ─────────────────────────────────────────
    roles_json = {"roles":[
        {"id":"aiops_engineer","label":"AIOps Engineer","color":"#1D4ED8"},
        {"id":"service_manager","label":"Service Manager","color":"#7C3AED"},
        {"id":"exec","label":"Executive","color":"#0E7C7B"},
    ], "navItems":[
        {"id":"dash","acc":"dash","label":"Command Centre","icon":"⬡"},
        {"id":"inc","acc":"inc","label":"Incidents","icon":"🔥"},
        {"id":"topology","acc":"topology","label":"Service Map","icon":"🗺"},
        {"id":"workflow","acc":"workflow","label":"Workflows","icon":"⚡"},
        {"id":"mim","acc":"mim","label":"MIM","icon":"🚨"},
        {"id":"req","acc":"req","label":"Service Requests","icon":"📋"},
        {"id":"agents","acc":"agents","label":"Agents & KB","icon":"🤖"},
        {"id":"reports","acc":"reports","label":"Reports","icon":"📊"},
        {"id":"transfer","acc":"transfer","label":"Transfer Readiness","icon":"🔄"},
        {"id":"chat","acc":"chat","label":"Ask ZOVA","icon":"💬"},
    ]}
    (out/"roles.json").write_text(json.dumps(roles_json, indent=2))
    print("  ✅ roles.json")

    # ── 4. metrics.json ───────────────────────────────────────
    def g(key, default=0):
        return safe_float(kpi.get(key, default), default)

    metrics_json = {"kpis":{
        "mttr_baseline_min":        g("mttr_baseline_min",3000),
        "mttr_current_min":         g("mttr_current_min",1200),
        "mttr_target_reduction_pct":g("mttr_target_reduction_pct",60),
        "auto_fix_baseline_pct":    g("auto_fix_baseline_pct",0),
        "auto_fix_rate_pct":        g("auto_fix_rate_pct",42),
        "auto_fix_target_pct":      g("auto_fix_target_pct",60),
        "alert_volume_per_day":     g("alert_volume_per_day",4000),
        "actionable_alerts_per_day":g("actionable_alerts_per_day",320),
        "alert_noise_red_pct":      g("alert_noise_red_pct",92),
        "csat_baseline":            g("csat_baseline",2.8),
        "csat_current":             g("csat_current",4.0),
        "nps_baseline":             g("nps_baseline",-40),
        "nps_current":              g("nps_current",20),
        "nps_target":               g("nps_target",50),
        "engineer_toil_baseline_pct":g("engineer_toil_baseline_pct",70),
        "engineer_toil_current_pct":g("engineer_toil_current_pct",28),
        "mttd_baseline_min":        g("mttd_baseline_min",45),
        "mttd_current_min":         g("mttd_current_min",4),
        "incidents_per_month":      g("incidents_per_month",450),
        "ticket_deflection_pct":    g("ticket_deflection_pct",48),
    }, "roi":{
        "llmCost":       g("llmCost",38),
        "netROI":        g("netROI",48322),
        "roiMultiplier": g("roiMultiplier",1272),
        "hoursSaved":    g("hoursSaved",266),
    }, "history": [
        {"month": f"M-{6-i}",
         "mttr": round(g("mttr_baseline_min",3000) * (1 - i*0.1)),
         "autoFix": round(i * g("auto_fix_rate_pct",42)/6),
         "csat": round(g("csat_baseline",2.8) + i*(g("csat_current",4.0)-g("csat_baseline",2.8))/6, 1)
        } for i in range(7)
    ]}
    (out/"metrics.json").write_text(json.dumps(metrics_json, indent=2))
    print("  ✅ metrics.json")

    # ── 5. applications.json ──────────────────────────────────
    domains = {}
    for row in apps:
        dn = row.get("domain_name","")
        if not dn: continue
        if dn not in domains:
            domains[dn] = {"name":dn,"id":re.sub(r"[^a-z0-9]","-",dn.lower()),
                           "status":"GREEN","uptime":99.5,"sessions":0,
                           "hosting":"Cloud","hostingColor":"#1D4ED8","trend":[],"apps":[]}
        trend_raw = row.get("trend_6m","99,99,99,99,99,99").split(",")
        trend = [safe_float(v,99) for v in trend_raw]
        u = safe_float(row.get("uptime_%","99.5"),99.5)
        domains[dn]["uptime"] = u
        domains[dn]["hosting"] = row.get("hosting","Cloud")
        domains[dn]["hostingColor"] = row.get("hosting_color","#CC0000")
        domains[dn]["trend"] = trend
        domains[dn]["sessions"] = (domains[dn]["sessions"] or 0) + safe_int(row.get("sessions","0"))

        app_status = row.get("status","GREEN").upper()
        if app_status == "RED" and domains[dn]["status"] != "RED":
            domains[dn]["status"] = "RED"
        elif app_status == "AMBER" and domains[dn]["status"] == "GREEN":
            domains[dn]["status"] = "AMBER"

        infra_vals = row.get("tech_stack","").split("·")
        domains[dn]["apps"].append({
            "id": re.sub(r"[^a-z0-9]","-",row.get("app_name","app").lower()),
            "name": row.get("app_name",""),
            "status": app_status,
            "tech": row.get("tech_stack",""),
            "perf": safe_int(row.get("perf_%","85")),
            "hosting": row.get("hosting","Cloud"),
            "infra": [
                {"c":"cpu","label":"CPU","val":f"{random.randint(30,75)}%","unit":"%","m":random.randint(30,75),"anom":app_status!="GREEN","detail":"Current CPU utilisation","h":app_status},
                {"c":"mem","label":"Memory","val":f"{random.randint(40,80)}%","unit":"%","m":random.randint(40,80),"anom":False,"detail":"Memory utilisation","h":"GREEN"},
            ]
        })

    apps_json = {"domains": list(domains.values())}
    (out/"applications.json").write_text(json.dumps(apps_json, indent=2))
    print("  ✅ applications.json")

    # ── 6. incidents.json ─────────────────────────────────────
    def build_intel(row):
        return {
            "cause": {
                "summary": row.get("notes",""),
                "confidence": 91,
                "signals": [
                    {"icon":"📊","t":f"ELK Stack · {row.get('detected_at','')}","v":f"Alert: {row.get('service_name','')} — {row.get('category','')} anomaly detected"},
                    {"icon":"🏛","t":f"Oracle OEM · {row.get('detected_at','')}","v":f"Threshold breach on {row.get('ci_reference','')}"},
                    {"icon":"🤖","t":"Alert Correlator","v":f"Correlated alerts → 1 incident ({row.get('priority','P2')})"},
                ],
                "agents": ["Alert Correlator","RCA Engine","Remediation Agent"]
            },
            "timeline": [
                {"t":"T+0s","event":"Alert Correlator detects signal from ELK"},
                {"t":"T+32s","event":"RCA Engine identifies root cause — 91% confidence"},
                {"t":"T+2m","event":"Remediation Agent executes pre-approved fix"},
                {"t":"T+5m","event":"Change Validator confirms recovery"},
                {"t":"T+6m","event":"ITSM ticket auto-closed"},
            ]
        }

    incidents_json = {"incidents":[{
        "id":     row.get("incident_id","INC-001"),
        "svc":    row.get("service_name",""),
        "pri":    row.get("priority","P2"),
        "status": row.get("status","Open"),
        "cat":    row.get("category",""),
        "hosting":row.get("hosting","Cloud"),
        "by":     row.get("assigned_to","ZeroOps"),
        "at":     row.get("detected_at",""),
        "ci":     row.get("ci_reference",""),
        "sla":    row.get("sla_status","MET"),
        "desc":   row.get("description",""),
        "notes":  row.get("notes",""),
        "chainId":row.get("chain_id",""),
        "appId":  re.sub(r"[^a-z0-9]","-",row.get("service_name","").lower()[:20]),
        "intel":  build_intel(row),
    } for row in incs if row.get("incident_id")]}
    (out/"incidents.json").write_text(json.dumps(incidents_json, indent=2))
    print("  ✅ incidents.json")

    # ── 7. agents.json ───────────────────────────────────────
    def build_anatomy(agent_name, tools_str, integrations_str):
        return {
            "trigger": f"Alert or incident routed from Alert Correlator requiring {agent_name} analysis",
            "process": f"Analyse signals → identify pattern → execute pre-approved action",
            "output":  f"Resolved incident with full audit trail in ITSM",
            "role":    f"{agent_name} — autonomous AI agent within Azure AI Foundry",
            "tasks":   [f"Ingest signals from {t.strip()}" for t in tools_str.split(",")[:3]] + ["Execute action","Validate outcome"],
            "contextUsed": [t.strip() for t in integrations_str.split(",")[:4]],
            "expectedOutput": f"Resolution executed — ITSM ticket updated and auto-closed",
            "decisionCriteria": "Risk score <25: auto-execute. 25-50: HiTL approval. >50: senior engineer.",
            "memoryType": "Episodic + pattern library",
            "llmReasoning": f"Claude Sonnet analyses {tools_str.split(',')[0].strip()} data to identify root cause and select pre-approved remediation"
        }

    agents_json = {"agents":[{
        "id":           re.sub(r"[^a-z0-9]","-",row.get("agent_name","agent").lower()),
        "name":         row.get("agent_name",""),
        "icon":         "🤖",
        "desc":         f"Autonomous AI agent — {row.get('agent_name','')}",
        "runs":         row.get("runs","0"),
        "success":      row.get("success_%","95%"),
        "tokens":       row.get("tokens_month","1M"),
        "costOK":       row.get("cost_per_ok","$0.01"),
        "human":        row.get("human_%","0%"),
        "status":       "Active",
        "model":        row.get("model","Claude Sonnet"),
        "provider":     "Azure AI Foundry",
        "ok":           safe_int(row.get("success_%","95").replace("%","")),
        "dur":          f"{row.get('avg_duration_s','5')}s",
        "tokensMonth":  row.get("tokens_month","1M"),
        "tools":        [t.strip() for t in row.get("tools","").split(",") if t.strip()],
        "integrations": [t.strip() for t in row.get("integrations","").split(",") if t.strip()],
        "drift":        {"dir":"stable","label":"Stable","delta":"0%"},
        "anatomy":      build_anatomy(
            row.get("agent_name",""), row.get("tools",""), row.get("integrations",""))
    } for row in agents if row.get("agent_name")]}
    (out/"agents.json").write_text(json.dumps(agents_json, indent=2))
    print("  ✅ agents.json")

    # ── 8. scenarios.json ─────────────────────────────────────
    def build_steps(scen):
        pillar = safe_int(scen.get("pillar","1"))
        auto_steps = safe_int(scen.get("automation","6/7").split("/")[0])
        total_steps = safe_int(scen.get("automation","6/7").split("/")[-1])
        steps = []
        base_steps = [
            ("Alert Correlator","Detect and correlate alerts from ELK","alert_corr","auto"),
            ("RCA Engine","Traverse dependency graph — root cause identified","rca","auto"),
            ("Log Analyzer","Deep log analysis — confirm root cause signal","log_anal","auto"),
            ("Remediation Agent","Prepare and execute pre-approved fix","remediation","auto"),
            ("Human Approval","Finance/Operations sign-off on fix","hitl","hitl"),
            ("Change Validator","Validate recovery across all affected systems","change_val","auto"),
            ("ITSM Connector","Auto-close ticket with full resolution evidence","xurrent","auto"),
        ]
        for i, (agent, label, aid, stype) in enumerate(base_steps[:total_steps]):
            is_hitl = (stype=="hitl" and pillar==2)
            steps.append({
                "id": f"step-{i+1}",
                "label": label,
                "icon": "👤" if is_hitl else "⚡",
                "type": "hitl" if is_hitl else "auto",
                "agent": agent,
                "summary": f"{agent}: {label}",
                "highlight": (i == 0 or i == total_steps-1),
                "duration": 4000,
                "events": [
                    {"msg": f"{agent} executing — {label}", "kind":"info", "t":500},
                    {"msg": "Completed successfully" if not is_hitl else "Awaiting human approval", "kind":"success" if not is_hitl else "warn", "t":3000},
                ]
            })
        return steps

    scenarios_json = {"scenarios":[{
        "id":          row.get("scenario_id","scenario-1"),
        "pillar":      safe_int(row.get("pillar","1")),
        "title":       row.get("title",""),
        "subtitle":    row.get("subtitle",""),
        "appId":       row.get("app_id",""),
        "riskScore":   safe_int(row.get("risk_score","20")),
        "pillarLabel": row.get("pillar_label","Pillar 1 — Sentinel"),
        "pillarColor": row.get("pillar_color","#0057A8"),
        "category":    row.get("category",""),
        "description": row.get("description",""),
        "blastRadius": row.get("blast_radius",""),
        "steps":       build_steps(row),
        "outcome": {
            "mttr":       row.get("mttr_after","18 min"),
            "automation": row.get("automation","6/7 steps automated"),
            "summary":    row.get("outcome_summary","Resolved autonomously"),
        }
    } for row in scens if row.get("scenario_id")]}
    (out/"scenarios.json").write_text(json.dumps(scenarios_json, indent=2))
    print("  ✅ scenarios.json")

    # ── 9. requests.json — inferred ──────────────────────────
    cats = {}
    for row in reqs:
        cat = row.get("category","General")
        if cat not in cats:
            cats[cat] = {"type":cat,"category":cat,"requests":[]}
        cats[cat]["requests"].append({
            "name":        row.get("request_name",""),
            "monthly":     safe_int(row.get("monthly_volume","10")),
            "baselineTime":row.get("baseline_time_hrs","24") + " hrs",
            "avgTime":     row.get("zeroops_time_hrs","4") + " hrs",
            "autoRate":    safe_int(row.get("auto_rate_%","50")),
            "sla":         row.get("sla","8 hrs"),
            "pillar":      row.get("pillar","Guardian"),
            "description": row.get("description",""),
        })

    total_today = sum(safe_int(r.get("monthly_volume","0")) for r in reqs) // 22
    auto_today  = sum(safe_int(r.get("monthly_volume","0")) * safe_int(r.get("auto_rate_%","0")) // 100 for r in reqs) // 22

    requests_json = {
        "catalogue": list(cats.values()),
        "stats": {
            "totalToday":    total_today,
            "autoFulfilled": auto_today,
            "autoRate":      round(auto_today/total_today*100) if total_today else 0,
            "openNow":       total_today * 3,
            "slaBreach":     total_today,
        }
    }
    (out/"requests.json").write_text(json.dumps(requests_json, indent=2))
    print("  ✅ requests.json")

    # ── 10. kb_articles.json — inferred from incidents ───────
    kb_articles = []
    for i, inc in enumerate(incidents_json["incidents"]):
        kb_articles.append({
            "id":        f"KB-{slug[:2].upper()}-{str(i+1).zfill(3)}",
            "title":     f"{inc['svc']} — Detection & Auto-Recovery",
            "domain":    inc.get("cat",""),
            "status":    "Published",
            "confidence":91,
            "lastUpdated":datetime.now().strftime("%d %b %Y"),
            "views":     random.randint(12,80),
            "body":      f"Auto-generated from incident {inc['id']}. {inc.get('desc','')}",
            "related_incidents": [inc["id"]],
            "related_agents":    inc.get("intel",{}).get("cause",{}).get("agents",[]),
            "sections": [
                {"heading":"Detection Pattern","body":f"ZeroOps detects this via ELK and Oracle OEM signals. Confidence: 91%."},
                {"heading":"Automated Resolution","body":f"Remediation Agent executes pre-approved fix. MTTR target: < 30 min."},
                {"heading":"Prevention","body":"Capacity Planner monitors leading indicators. Alert threshold configured in ELK."},
            ]
        })
    (out/"kb_articles.json").write_text(json.dumps({"articles":kb_articles}, indent=2))
    print("  ✅ kb_articles.json  (inferred from incidents)")

    # ── 11. silent_ops.json — inferred from scenarios ────────
    silent_ops = []
    times = ["05:14","05:32","06:40","07:18","08:03","08:47","09:22","10:05","11:30","12:14"]
    outcomes = ["AUTO-RESOLVED","HEALED","COMPLETED","AUTO-RESOLVED","HEALED"]
    for i, scen in enumerate(scenarios_json["scenarios"]):
        silent_ops.append({
            "id":       f"SILOPS-{str(i+1).zfill(4)}",
            "ts":       times[i % len(times)],
            "svc":      scen["title"][:40],
            "outcome":  outcomes[i % len(outcomes)],
            "pillar":   scen["pillar"],
            "mttr":     scen["outcome"]["mttr"],
            "saved":    f"~ {scen.get('subtitle','').split('→')[0].split('·')[-1].strip()} avoided",
            "action":   scen["outcome"]["summary"],
            "agents":   ["Alert Correlator","RCA Engine","Remediation Agent"],
        })
    # Add some padding entries
    padding = [
        {"id":"SILOPS-P001","ts":"03:41","svc":"Alert Noise Suppression","outcome":"COMPLETED","pillar":1,"mttr":"0.3s","saved":"143 duplicate alerts","action":"Alert Correlator suppressed 143 correlated alerts → 1 incident","agents":["Alert Correlator"]},
        {"id":"SILOPS-P002","ts":"02:14","svc":"Capacity Threshold Monitor","outcome":"HEALED","pillar":1,"mttr":"2 min","saved":"Potential disk full event","action":"Capacity Planner detected trend breach — auto-cleanup executed","agents":["Capacity Planner","Remediation Agent"]},
        {"id":"SILOPS-P003","ts":"01:08","svc":"CMDB Drift Detection","outcome":"COMPLETED","pillar":1,"mttr":"4 min","saved":"Configuration inconsistency","action":"Change Validator detected and corrected 3 CMDB drift items","agents":["Change Validator"]},
    ]
    silent_ops = padding + silent_ops
    (out/"silent_ops.json").write_text(json.dumps({"log":silent_ops}, indent=2))
    print("  ✅ silent_ops.json   (inferred from scenarios)")

    # ── 12. mim_teams.json — from MIM sheet ─────────────────
    sme_roster = {inc["id"]: [
        {"name":r.get("name",""),"role":r.get("role",""),"team":r.get("team",""),"status":"Available"}
        for r in mim
    ] for inc in incidents_json["incidents"]}

    mim_json = {
        "smeRoster":    sme_roster,
        "timeline":     {inc["id"]: [] for inc in incidents_json["incidents"]},
        "warRoomChat":  {inc["id"]: [] for inc in incidents_json["incidents"]},
        "vendorEscalations": {inc["id"]: [] for inc in incidents_json["incidents"]},
        "broadcastTemplates": [
            {"id":"b1","label":"P1 Declared","text":f"🚨 P1 DECLARED — {cname} Production Impact. ZeroOps MIM activated. War room open. SMEs assemble."},
            {"id":"b2","label":"Update","text":"🔄 UPDATE — Root cause identified. Remediation in progress. ETA 15 min."},
            {"id":"b3","label":"Resolved","text":"✅ RESOLVED — Incident resolved. ZeroOps autonomous recovery completed. ITSM ticket closed."},
        ]
    }
    (out/"mim_teams.json").write_text(json.dumps(mim_json, indent=2))
    print("  ✅ mim_teams.json")

    # ── 13. chat_suggestions.json — inferred from metrics ────
    auto_fix  = safe_int(kpi.get("auto_fix_rate_pct","42"))
    mttr_curr = safe_int(kpi.get("mttr_current_min","1200"))
    mttr_hrs  = round(mttr_curr/60)
    noise     = safe_int(kpi.get("alert_noise_red_pct","92"))
    vol       = safe_int(kpi.get("alert_volume_per_day","4000"))
    actable   = safe_int(kpi.get("actionable_alerts_per_day","320"))

    chat_json = {"suggestions":[
        {"icon":"🔥","text":"What P1 incidents are active right now?"},
        {"icon":"⚡","text":f"How did ZeroOps resolve today's incidents autonomously?"},
        {"icon":"📊","text":f"Our MTTR is now {mttr_hrs}h — what drove the improvement?"},
        {"icon":"🤖","text":f"{auto_fix}% auto-fix rate — which agents are performing best?"},
        {"icon":"🔇","text":f"{noise}% alert noise reduction — how does the Alert Correlator work?"},
        {"icon":"💰","text":"What is the ROI on ZeroOps this month?"},
        {"icon":"🗺","text":"Which applications are currently degraded?"},
        {"icon":"📋","text":"Show me the Silent Operations Centre log"},
        {"icon":"🔄","text":"What is the Transfer Readiness status?"},
        {"icon":"🎫","text":"How does the ITSM integrate with ZeroOps?"},
    ]}
    (out/"chat_suggestions.json").write_text(json.dumps(chat_json, indent=2))
    print("  ✅ chat_suggestions.json (inferred from metrics)")

    print(f"\n🎉 Done! 13 JSON files generated in: {out}")
    print(f"   Set ACTIVE_CUSTOMER = '{slug}' in loader.js to activate")


# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="ZeroOps Customer JSON Generator")
    parser.add_argument("--workbook", type=str, required=True,
                        help="Path to filled-in ZeroOps contextualisation workbook (.xlsx)")
    args = parser.parse_args()

    if not Path(args.workbook).exists():
        print(f"❌ File not found: {args.workbook}")
        sys.exit(1)

    generate_jsons(args.workbook)

